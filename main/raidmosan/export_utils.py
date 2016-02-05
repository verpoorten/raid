from django.http import HttpResponse

import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.cell import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Fill, Color, Style, PatternFill
from openpyxl.worksheet import Worksheet, ColumnDimension, RowDimension
from raidmosan.models import Inscription
from django.utils.dateformat import DateFormat
from django.utils.formats import get_format
from django.utils.translation import ugettext_lazy as _

HEADER = [str('Equipe'),
          str('Catégorie'),
          str('Participant 1'),
          str('Participant 2'),
          str('T-shirt 1'),
          str('T-shirt 2'),
          str('Email responsable'),
          str('Téléphone responsable'),
          str('Reglement accepté'),
          str('Raid trophy'),]

def export_xls(modeladmin, request, inscriptions):
    print('export_xls')
    wb = Workbook()
    ws = wb.active
    ws.append(HEADER)
    ws.column_dimensions['A'].width=30
    ws.column_dimensions['C'].width=30
    ws.column_dimensions['D'].width=30
    ws.column_dimensions['G'].width=30
    ws.column_dimensions['H'].width=30

    cptr=1
    for inscription in inscriptions:

        ws.append([inscription.nom_equipe,
                   inscription.categorie,
                   inscription.nom_participant_1,
                   inscription.nom_participant_2,
                   inscription.taille_t_shirt_1,
                   inscription.taille_t_shirt_2,
                   inscription.responsable_email,
                   inscription.responsable_telephone,
                   inscription.reglement_ok,
                   inscription.challenge_raid_trophy])

        cptr = cptr+1

    response = HttpResponse(content=save_virtual_workbook(wb))
    response['Content-Disposition'] = 'attachment; filename=inscriptions.xlsx'
    response['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
